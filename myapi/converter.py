from html.parser import HTMLParser
import math
import copy
import re
import os

from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, rows_from_range

from myapi.parsers import TableParser
import myapi.tests.onedrive_test as onedrive


class Converter():
    '''
    Converts an html table into an excel file

    compress: tries to minimize the number of rows and columns used, removing empty rows and columns
    match_style: retains style of html table, e.g. bold font
    '''
    def __init__(self, html, compress=True, match_style=True, save_path=os.environ.get('TABLE_PATH')):
        self.match_style = match_style
        self.save_path = save_path
        parser = TableParser(html)
        if compress:
            comp = Compresser(parser.parsed_rows)
            self.rows = comp.rows
        else:
            self.rows = parser.parsed_rows

        self.wb = Workbook()
        self.ws = self.wb.active

    def upload(self):
        onedrive.upload()

    def convert(self):
        cells_to_merge = {}
        col_widths = []
        row_heights = []
        for y,row in enumerate(self.rows):
            for x, cell in enumerate(row):
                self.format_cell(x+1, y+1, cell)

                cell_width = cell.attrs['colspan']
                normalized_val_width = self.val_width(cell.data)/cell_width + 2
                max_col_width = self.max_col_width(x)
                width = min(normalized_val_width, max_col_width)
                if y == 0:
                    col_widths.append(width)
                else:
                    col_widths[x] = max(col_widths[x], width)

                val_height = self.val_height(cell.data, cell_width, x)
                cell_height = cell.attrs['rowspan']
                normalized_val_height = val_height/cell_height
                if x == 0:
                    row_heights.append(normalized_val_height)
                else:
                    row_heights[y] = max(row_heights[y], normalized_val_height)

                i = (cell.col, cell.row)
                if cell_width > 1 or cell_height > 1:
                    if i not in cells_to_merge:
                        cells_to_merge[i] = [(x+1, y+1)]
                    else:
                        cells_to_merge[i].append((x+1, y+1))

        self.set_col_widths(col_widths)
        self.set_row_heights(row_heights)
        self.merge_cells(cells_to_merge)
        self.wb.save(self.save_path)

    def set_col_widths(self, col_widths):
        first_col_width = col_widths[0]
        other_cols_width = max(col_widths[1:]) if col_widths[1:] else 0
        for i in range(len(col_widths)):
            width = first_col_width if i == 0 else other_cols_width
            col_letter = get_column_letter(i+1)
            self.ws.column_dimensions[col_letter].width = width

    def set_row_heights(self, row_heights):
        for i, height in enumerate(row_heights):
            self.ws.row_dimensions[i+1].height = height

    def merge_cells(self, cells):
        for bounds in cells.values():
            left = bounds[0][0]
            top = bounds[0][1]
            right = bounds[-1][0]
            bottom = bounds[-1][1]
            self.ws.merge_cells(start_column=left, start_row=top,
                                    end_column=right, end_row=bottom)

    def format_cell(self, x, y, parsed_cell):
        val = parsed_cell.data
        if num_type := self.numeric_type(val):
            val, is_int = self.format_number(val)
        cell = self.ws.cell(row=y, column=x, value=val)
        if num_type:
            self.format_cell_number(cell, num_type, is_int)
        else:
            self.set_wrap_text(cell)

        if self.match_style:
            self.set_style(cell, parsed_cell.style)

    def format_cell_number(self, cell, num_type, is_int):
        if is_int:
            base = '#,###'
        else:
            base = '#,##0.00'

        if num_type == 'year':
            cell.number_format = '###0'
        elif num_type == 'number':
            cell.number_format = f'{base}_);({base})_);@_)'
        elif num_type == 'dollar':
            cell.number_format = f'_)$* {base}_);_)$* ({base})_);_)$* _)@_)'
        elif num_type == 'percent':
            cell.number_format = f'{base}%;({base})%;@\%'

    def format_number(self, val):
        is_negative = True if '(' in val and ')' in val else False
        is_percent = True if '%' in val else False

        for c in ['(',')','$','%',',']:
            val = val.replace(c,'')
        if val.strip() == '—':
            return val, True

        num = float(val) if '.' in val else int(val)
        is_int = isinstance(num, int)
        num = -num if is_negative else num
        num = num/100 if is_percent else num

        return num, is_int

    def numeric_type(self, val):
        dollar = re.compile(r'\$\s*[\(\-]?\s*[0-9\—]+[^A-Za-z\s]*[0-9]*\s*\)?$')
        number = re.compile(r'[\(\-]?\s*[0-9\—]+[^A-Za-z\s\-%*/]*[0-9]*\s*\)?$')
        percent = re.compile(r'[\(\-]?\s*[0-9\—]+[^A-Za-z\s]*[0-9]*\s*\)?\s*%\s*\)?$')
        if number.match(val):
            if len(val.strip()) == 4 and not any(c in val for c in [',','.','(',')']):
                return 'year'
            return 'number'
        elif dollar.match(val):
            return 'dollar'
        elif percent.match(val):
            return 'percent'
        return None

    def val_width(self, val):
        val_str = self.to_string(val)
        lines = val_str.split('\n')
        return max(len(line) for line in lines)

    def val_height(self, val, cell_width, col):
        val_str = self.to_string(val)
        lines = val_str.split('\n')
        height = 0
        for line in lines:
            height += 15*math.ceil(len(line)/self.max_col_width(col)/cell_width)
        return height

    def max_col_width(self, col):
        if col == 0:
            return 80
        else:
            return 25

    def to_string(self, val):
        if not val:
            return ''
        elif isinstance(val, float):
            return f'{val:.2f}'
        return str(val)

    def set_style(self, cell, style):
        if 'font_weight' in style:
            weight = style['font_weight']
            if weight == 'bold' or (weight.isnumeric() and int(weight) > 500):
                self.set_bold(cell)
        if 'text_align' in style:
            self.set_alignment(cell, style['text_align'])
        if 'vertical_align' in style:
            self.set_alignment(cell, style['vertical_align'])

    def set_alignment(self, cell, align):
        alignment = copy.copy(cell.alignment)
        if align in ['left', 'right', 'center']:
            alignment.horizontal = align
        elif align in ['bottom', 'top']:
            alignment.vertical = align
        cell.alignment = alignment

    def set_bold(self, cell):
        font = copy.copy(cell.font)
        font.bold = True
        cell.font = font

    def set_wrap_text(self, cell):
        alignment = copy.copy(cell.alignment)
        alignment.wrapText = True
        cell.alignment = alignment

class Compresser():
    '''
    Attempts to minimize size of table array by:
        Removing empty columns and rows
        Removing duplicate columns and rows
        Moving unit symbols (e.g. $, %) into the same cell as the numerical value
    '''
    def __init__(self, parsed_rows):
        self.rows = parsed_rows
        self.compress()

    def compress(self):
        self.remove_empty_rows()
        self.match_cols()
        self.remove_blank_columns()
        self.remove_duplicate_columns()
        self.move_symbols()
        self.remove_duplicate_rows()
        self.remove_blank_rows()

    def remove_blank_columns(self):
        removable_columns = []
        for i in range(len(self.rows[0])):
            if self.is_removable_column(i):
                self.remove_column(i)
                self.remove_blank_columns()
                return None
        return None

    def remove_blank_rows(self):
        removable_columns = []
        for i in range(len(self.rows)):
            if self.is_removable_row(i):
                self.remove_row(i)
                self.remove_blank_rows()
                return None
        return None

    def remove_duplicate_columns(self):
        removable_columns = []
        for x in range(len(self.rows[0])-1):
            col0 = [row[x].data for row in self.rows]
            col1 = [row[x+1].data for row in self.rows]
            if col0 == col1 and x not in removable_columns:
                removable_columns.append(x)
        self.remove_columns(removable_columns)

    def remove_duplicate_rows(self):
        removable_rows = []
        for y in range(len(self.rows)-1):
            row0 = self.rows[y]
            row1 = self.rows[y+1]
            if row0 == row1 and y not in removable_rows:
                removable_rows.append(y)
        self.remove_rows(removable_rows)

    def remove_empty_rows(self):
        rows_to_remove = []
        for y, row in enumerate(self.rows):
            if not [cell for cell in row if cell.data]:
                rows_to_remove.append(y)

        self.remove_rows(rows_to_remove)
        if not self.rows:
            raise Exception('Empty table')

    def move_symbols(self):
        removable_columns = set()
        for y,row in enumerate(self.rows):
            for x, cell in enumerate(row):
                data = cell.data
                if self.symbol_side(data) == 'left':
                    if x+1 < len(self.rows[y]) and self.rows[y][x+1].data and not self.symbol_side(self.rows[y][x+1].data):
                        self.rows[y][x+1].data = f'{data}{self.rows[y][x+1].data}'
                        removable_columns.add(x)
                elif self.symbol_side(data) == 'right':
                    if x > 0 and self.rows[y][x-1].data and not self.symbol_side(self.rows[y][x-1].data):
                        self.rows[y][x-1].data = f'{self.rows[y][x-1].data}{data}'
                        removable_columns.add(x)
        self.remove_columns(removable_columns)

    def match_cols(self):
        max_col = max(len(row) for row in self.rows)
        for row in self.rows:
            last_cell = row[-1]
            while len(row) < max_col:
                row.append(last_cell)
                last_cell.attrs['colspan'] += 1

    def symbol_side(self, data):
        if self.is_left_symbol(data):
            return 'left'
        elif self.is_right_symbol(data):
            return 'right'
        return None

    def is_left_symbol(self, data):
        return data in ['$', '(']

    def is_right_symbol(self, data):
        return data in ['%', ')']
        #footnote = re.compile(r'[\[\(][A-Za-z0-9][\]\)]$')
        #if footnote.match(data):
        #    return True

    def is_removable_column(self, col):
        mid = [row[col].data for row in self.rows]
        left = [row[col-1].data for row in self.rows] if col > 0 else ['']*len(mid)
        right = [row[col+1].data for row in self.rows] if col < len(self.rows[0])-1 else ['']*len(mid)
        return not [el for i,el in enumerate(mid) if el and el != left[i] and el != right[i]]

    def is_removable_row(self, row):
        mid = [cell.data for cell in self.rows[row]]
        upper = [cell.data for cell in self.rows[row-1]] if row > 0 else ['']*len(mid)
        lower = [cell.data for cell in self.rows[row+1]] if row < len(self.rows)-1 else ['']*len(mid)
        return not [el for i,el in enumerate(mid) if el and el != upper[i] and el != lower[i]]

    def remove_rows(self, rows):
        rows.sort()
        for i,row in enumerate(rows):
            row -= i
            self.remove_row(row)

    def remove_columns(self, cols):
        cols = list(cols)
        cols.sort()
        for i,col in enumerate(cols):
            col -= i
            self.remove_column(col)

    def remove_row(self, row):
        for cell in set(self.rows[row]):
            cell.attrs['rowspan'] -= 1
        self.rows.pop(row)

    def remove_column(self, col):
        skip = []
        for row in self.rows:
            cell = row[col]
            if cell not in skip:
                cell.attrs['colspan'] -= 1
                if cell.attrs['rowspan'] > 1:
                    skip.append(cell)
            row.pop(col)
